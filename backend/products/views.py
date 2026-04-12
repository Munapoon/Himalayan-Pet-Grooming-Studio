from django.shortcuts import render, redirect, get_object_or_404
from django.conf import settings
from django.contrib.auth.decorators import login_required
from django.contrib import messages
from django.urls import reverse
from django.core.paginator import Paginator
from django.db.models import Q
import requests as http_requests
from django.template.loader import render_to_string
from django.utils.html import strip_tags
import decimal
import uuid
from django.utils import timezone

from accounts.decorators import admin_required
from .models import ProductCategory, Product, Cart, Order, OrderItem, Payment, Sale
from .forms import ProductCategoryForm, ProductForm


# Product Category Views
@login_required
@admin_required
def category_list(request):
    from django.core.paginator import Paginator
    
    categories = ProductCategory.objects.all().order_by('-created_at')
    
    # Pagination - 10 categories per page
    paginator = Paginator(categories, 10)
    page_number = request.GET.get('page')
    page_obj = paginator.get_page(page_number)
    
    return render(request, 'products/category_list.html', {'page_obj': page_obj})


@login_required
@admin_required
def category_create(request):
    if request.method == 'POST':
        form = ProductCategoryForm(request.POST)
        if form.is_valid():
            form.save()
            messages.success(request, 'Category created successfully!')
            return redirect('products:category_list')
    else:
        form = ProductCategoryForm()
    return render(request, 'products/category_form.html', {'form': form, 'action': 'Create'})


@login_required
@admin_required
def category_update(request, pk):
    category = get_object_or_404(ProductCategory, pk=pk)
    if request.method == 'POST':
        form = ProductCategoryForm(request.POST, instance=category)
        if form.is_valid():
            form.save()
            messages.success(request, 'Category updated successfully!')
            return redirect('products:category_list')
    else:
        form = ProductCategoryForm(instance=category)
    return render(request, 'products/category_form.html', {'form': form, 'action': 'Update', 'category': category})


@login_required
@admin_required
def category_delete(request, pk):
    category = get_object_or_404(ProductCategory, pk=pk)
    if request.method == 'POST':
        # Block deletion if category has active products
        active_products = category.products.filter(is_active=True)
        active_count = active_products.count()
        if active_count > 0:
            messages.error(
                request,
                f'Cannot delete "{category.name}" category. '
                f'It has {active_count} active product(s) assigned to it. '
                f'Please deactivate or reassign those products before deleting this category.'
            )
            return redirect('products:category_list')
        category.delete()
        messages.success(request, f'Category "{category.name}" deleted successfully!')
        return redirect('products:category_list')
    return render(request, 'products/category_confirm_delete.html', {'category': category})


# Product Views
def product_list(request):
    from django.core.paginator import Paginator
    from django.db.models import Q
    
    # Get search query, category filter, price range, and sort option
    search_query = request.GET.get('search') or request.GET.get('q', '')
    category_id = request.GET.get('category', '')
    min_price = request.GET.get('min_price', '')
    max_price = request.GET.get('max_price', '')
    sort_by = request.GET.get('sort', 'newest')
    
    if request.user.is_authenticated and request.user.is_admin_user():
        products = Product.objects.all()
        template_name = 'products/product_list_admin.html'
    else:
        products = Product.objects.filter(is_active=True)
        template_name = 'products/product_list.html'
    
    # Apply search filter
    if search_query:
        products = products.filter(
            Q(name__icontains=search_query) |
            Q(description__icontains=search_query) |
            Q(category__name__icontains=search_query)
        )
    
    # Apply category filter
    if category_id:
        products = products.filter(category_id=category_id)

    # Apply price filters
    if min_price:
        try:
            products = products.filter(price__gte=float(min_price))
        except ValueError:
            pass
            
    if max_price:
        try:
            products = products.filter(price__lte=float(max_price))
        except ValueError:
            pass
    
    # Apply sorting
    if sort_by == 'price_low':
        products = products.order_by('price')
    elif sort_by == 'price_high':
        products = products.order_by('-price')
    elif sort_by == 'rating':
        products = products.order_by('-rating', '-created_at')
    else:  # newest (default)
        products = products.order_by('-created_at')
    
    # Pagination - 10 products per page
    paginator = Paginator(products, 10)
    page_number = request.GET.get('page')
    page_obj = paginator.get_page(page_number)
    
    # Get all categories for filter
    categories = ProductCategory.objects.all()
    
    context = {
        'page_obj': page_obj,
        'search_query': search_query,
        'category_id': category_id,
        'min_price': min_price,
        'max_price': max_price,
        'sort_by': sort_by,
        'categories': categories,
    }
    
    return render(request, template_name, context)


@login_required
@admin_required
def product_create(request):
    if request.method == 'POST':
        form = ProductForm(request.POST, request.FILES)
        if form.is_valid():
            form.save()
            messages.success(request, 'Product created successfully!')
            return redirect('products:product_list')
    else:
        form = ProductForm()
    return render(request, 'products/product_form.html', {'form': form, 'action': 'Create'})


def product_detail(request, pk):
    from .models import Review, Sale, OrderItem
    product = get_object_or_404(Product, pk=pk)
    
    # Check if user has purchased this product
    # Check BOTH: Sale records (admin) AND actual OrderItems from user orders
    has_purchased = False
    if request.user.is_authenticated and not request.user.is_admin_user():
        via_sale = Sale.objects.filter(
            customer=request.user,
            product=product
        ).exists()
        via_order = OrderItem.objects.filter(
            order__user=request.user,
            product=product
        ).exists()
        has_purchased = via_sale or via_order
    
    # Get user's existing review if any
    user_review = None
    if request.user.is_authenticated and not request.user.is_admin_user():
        user_review = Review.objects.filter(product=product, user=request.user).first()
    
    # Get all reviews for this product (Approved only, or owner's own)
    reviews_qs = Review.objects.filter(product=product).select_related('user')
    if not request.user.is_authenticated or not request.user.is_admin_user():
        # Regular users see approved ones or their own
        if request.user.is_authenticated:
            reviews = reviews_qs.filter(Q(is_approved=True) | Q(user=request.user))
        else:
            reviews = reviews_qs.filter(is_approved=True)
    else:
        # Admins see everything
        reviews = reviews_qs
    
    # Simple Recommendation System
    from appointments.models import Service
    recommended_service = None
    
    product_text = f"{product.name} {product.category.name}".lower()
    
    if "nail" in product_text or "clipper" in product_text or "trimmer" in product_text:
        recommended_service = Service.objects.filter(is_active=True, slug__icontains='nail').first()
    elif "shampoo" in product_text or "conditioner" in product_text or "bath" in product_text or "soap" in product_text:
        recommended_service = Service.objects.filter(is_active=True, slug__icontains='bath').first()
    elif "scissor" in product_text or "brush" in product_text or "comb" in product_text or "hair" in product_text:
        recommended_service = Service.objects.filter(is_active=True, slug__icontains='haircut').first()
    elif "oil" in product_text or "spa" in product_text or "massage" in product_text:
        recommended_service = Service.objects.filter(is_active=True, slug__icontains='spa').first()
        
    if not recommended_service:
        # Fallback to a deterministic service based on product ID so it doesn't change on refresh
        active_services = list(Service.objects.filter(is_active=True))
        if active_services:
            index = product.id % len(active_services)
            recommended_service = active_services[index]


    context = {
        'product': product,
        'has_purchased': has_purchased,
        'user_review': user_review,
        'reviews': reviews,
        'recommended_service': recommended_service,
    }
    
    if request.user.is_authenticated and request.user.is_admin_user():
        template_name = 'products/product_detail_admin.html'
    else:
        template_name = 'products/product_detail.html'
    return render(request, template_name, context)


@login_required
@admin_required
def product_update(request, pk):
    product = get_object_or_404(Product, pk=pk)
    if request.method == 'POST':
        form = ProductForm(request.POST, request.FILES, instance=product)
        if form.is_valid():
            form.save()
            messages.success(request, 'Product updated successfully!')
            return redirect('products:product_list')
    else:
        form = ProductForm(instance=product)
    return render(request, 'products/product_form.html', {'form': form, 'action': 'Update', 'product': product})


@login_required
@admin_required
def product_delete(request, pk):
    product = get_object_or_404(Product, pk=pk)
    if request.method == 'POST':
        product.delete()
        messages.success(request, 'Product deleted successfully!')
        return redirect('products:product_list')
    return render(request, 'products/product_confirm_delete.html', {'product': product})


# Cart Views
@login_required
def cart_view(request):
    """View cart items"""
    cart_items = Cart.objects.filter(user=request.user).select_related('product')
    
    total = sum(item.subtotal for item in cart_items)
    
    context = {
        'cart_items': cart_items,
        'total': total,
    }
    return render(request, 'products/cart.html', context)


@login_required
def add_to_cart(request, product_id):
    """Add product to cart"""
    product = get_object_or_404(Product, id=product_id)
    
    if not product.in_stock:
        messages.error(request, 'This product is currently out of stock.')
        return redirect('products:product_detail', pk=product_id)
    
    quantity = int(request.POST.get('quantity', 1))
    size = request.POST.get('size', None)
    redirect_to = request.POST.get('redirect_to', 'cart')
    
    # Check if item already in cart with same size
    cart_item, created = Cart.objects.get_or_create(
        user=request.user,
        product=product,
        size=size,
        defaults={'quantity': quantity}
    )
    
    if not created:
        # Update quantity if item already exists
        cart_item.quantity += quantity
        cart_item.save()
        messages.success(request, f'Updated {product.name} quantity in cart.')
    else:
        messages.success(request, f'{product.name} added to cart successfully!')
    
    # Redirect based on the redirect_to parameter
    if redirect_to == 'product':
        return redirect('products:product_detail', pk=product_id)
    else:
        return redirect('products:cart')


@login_required
def update_cart(request, cart_id):
    """Update cart item quantity"""
    cart_item = get_object_or_404(Cart, id=cart_id, user=request.user)
    
    if request.method == 'POST':
        quantity = int(request.POST.get('quantity', 1))
        
        if quantity > 0:
            if quantity <= cart_item.product.stock_quantity:
                cart_item.quantity = quantity
                cart_item.save()
                messages.success(request, 'Cart updated successfully!')
            else:
                messages.error(request, f'Only {cart_item.product.stock_quantity} items available in stock.')
        else:
            cart_item.delete()
            messages.success(request, 'Item removed from cart.')
    
    return redirect('products:cart')


@login_required
def remove_from_cart(request, cart_id):
    """Remove item from cart"""
    cart_item = get_object_or_404(Cart, id=cart_id, user=request.user)
    product_name = cart_item.product.name
    cart_item.delete()
    messages.success(request, f'{product_name} removed from cart.')
    return redirect('products:cart')


@login_required
def clear_cart(request):
    """Clear all items from cart"""
    Cart.objects.filter(user=request.user).delete()
    messages.success(request, 'Cart cleared successfully!')
    return redirect('products:cart')


@login_required
def add_review(request, pk):
    """Add or update a product review — only for users who purchased the product"""
    from .models import Review, Sale, OrderItem
    from .forms import ReviewForm
    
    product = get_object_or_404(Product, pk=pk)
    
    # Check if user has purchased this product via Sale OR OrderItem
    via_sale = Sale.objects.filter(
        customer=request.user,
        product=product
    ).exists()
    via_order = OrderItem.objects.filter(
        order__user=request.user,
        product=product
    ).exists()
    has_purchased = via_sale or via_order
    
    if not has_purchased:
        messages.error(request, 'You can only review products you have purchased.')
        return redirect('products:product_detail', pk=pk)
    
    # Check if user already has a review
    existing_review = Review.objects.filter(product=product, user=request.user).first()
    
    if request.method == 'POST':
        if existing_review:
            form = ReviewForm(request.POST, instance=existing_review)
        else:
            form = ReviewForm(request.POST)
        
        if form.is_valid():
            review = form.save(commit=False)
            review.product = product
            review.user = request.user
            review.save()
            messages.success(request, 'Your review has been submitted successfully!')
            return redirect('products:product_detail', pk=pk)
    else:
        if existing_review:
            form = ReviewForm(instance=existing_review)
        else:
            form = ReviewForm()
    
    context = {
        'form': form,
        'product': product,
        'existing_review': existing_review,
    }
    return render(request, 'products/review_form.html', context)


@login_required
def delete_review(request, pk):
    """Delete a product review"""
    from .models import Review
    
    review = get_object_or_404(Review, pk=pk)
    product_id = review.product.id
    
    # Check if user owns this review
    if review.user != request.user:
        messages.error(request, 'You can only delete your own reviews.')
        return redirect('products:product_detail', pk=product_id)
    
    if request.method == 'POST':
        review.delete()
        messages.success(request, 'Your review has been deleted.')
        return redirect('products:product_detail', pk=product_id)
    
    return render(request, 'products/review_confirm_delete.html', {'review': review})


# Checkout & Order Views
@login_required
def checkout(request):
    """Checkout page with payment options"""
    from .models import Order, OrderItem
    
    # Get selected item IDs from POST request
    selected_item_ids = request.POST.getlist('selected_items')
    
    if selected_item_ids:
        # Filter cart items by selected IDs
        cart_items = Cart.objects.filter(
            user=request.user, 
            id__in=selected_item_ids
        ).select_related('product')
        # Store selected IDs in session for later use
        request.session['selected_cart_items'] = selected_item_ids
    elif 'selected_cart_items' in request.session:
        # Use stored selection from session
        cart_items = Cart.objects.filter(
            user=request.user,
            id__in=request.session['selected_cart_items']
        ).select_related('product')
    else:
        # No selection - use all cart items
        cart_items = Cart.objects.filter(user=request.user).select_related('product')
    
    if not cart_items.exists():
        messages.warning(request, 'Your cart is empty or no items selected.')
        return redirect('products:cart')
    
    # Check stock availability
    for item in cart_items:
        if item.quantity > item.product.stock_quantity:
            messages.error(request, f'{item.product.name} has insufficient stock.')
            return redirect('products:cart')
    
    total = sum(item.subtotal for item in cart_items)
    
    if request.method == 'POST' and 'payment_method' in request.POST:
        payment_method = request.POST.get('payment_method')
        shipping_address = request.POST.get('shipping_address')
        phone_number = request.POST.get('phone_number')
        notes = request.POST.get('notes', '')
        
        if not shipping_address or not phone_number:
            messages.error(request, 'Please provide shipping address and phone number.')
            return render(request, 'products/checkout.html', {
                'cart_items': cart_items,
                'total': total,
            })
        
        # Create order
        order = Order.objects.create(
            user=request.user,
            total_amount=total,
            payment_method=payment_method,
            shipping_address=shipping_address,
            phone_number=phone_number,
            notes=notes,
        )
        
        # Create order items and update stock
        for item in cart_items:
            OrderItem.objects.create(
                order=order,
                product=item.product,
                size=item.size,
                quantity=item.quantity,
                unit_price=item.product.price,
            )
            
            # Update product stock
            item.product.stock_quantity -= item.quantity
            item.product.save()
            
            # Create Sale record for purchase verification (for reviews)
            Sale.objects.create(
                product=item.product,
                customer=request.user,
                quantity=item.quantity,
                unit_price=item.product.price,
                payment_method='online' if payment_method == 'khalti' else 'cash',
            )
        
        # Clear only the selected cart items
        cart_items.delete()
        
        # Clear session data
        if 'selected_cart_items' in request.session:
            del request.session['selected_cart_items']
        
        if payment_method == 'cod':
            # Cash on Delivery - Order is pending
            order.payment_status = 'pending'
            order.save()
            
            from .models import Payment
            Payment.objects.create(
                order=order,
                user=request.user,
                transaction_id=f"COD-{order.order_number}",
                payment_method='cod',
                amount=order.total_amount,
                status='pending',
            )
            
            # Send email confirmation
            from django.core.mail import send_mail
            
            subject = f"Order #{order.order_number} Placed"
            
            context = {
                'user_name': request.user.get_full_name() or request.user.username,
                'order': order,
                'site_url': request.build_absolute_uri('/')[:-1]
            }
            
            html_message = render_to_string('emails/order_confirmation.html', context)
            plain_message = strip_tags(html_message)
            
            try:
                send_mail(
                    subject, 
                    plain_message, 
                    settings.DEFAULT_FROM_EMAIL, 
                    [request.user.email], 
                    html_message=html_message,
                    fail_silently=True
                )
            except Exception as e:
                print(f"DEBUG: Email confirmation error (COD): {e}")
                pass
            
            messages.success(request, f'Order placed successfully! Order Number: {order.order_number}')
            return redirect('products:order_detail', order_number=order.order_number)
        
        elif payment_method == 'khalti':
            # Khalti payment - Initiate KPG-2
            import requests as http_requests
            from django.urls import reverse
            
            order.payment_status = 'pending'
            order.save()
            
            url = f"{settings.KHALTI_BASE_URL}epayment/initiate/"
            return_url = request.build_absolute_uri(reverse('products:khalti_verify'))
            
            # Build purchase name from product items
            product_names = ", ".join([item.product.name for item in order.items.all()])
            if len(product_names) > 80:
                product_names = product_names[:77] + "..."
            
            payload = {
                "return_url": return_url,
                "website_url": request.build_absolute_uri('/'),
                "amount": int(order.total_amount * 100),
                "purchase_order_id": f"order-{order.order_number}",
                "purchase_order_name": f"Order #{order.order_number}: {product_names}",
                "customer_info": {
                    "name": request.user.get_full_name() or request.user.username,
                    "email": request.user.email or "customer@example.com",
                    "phone": phone_number
                }
            }
            
            headers = {
                'Authorization': f"Key {settings.KHALTI_SECRET_KEY}",
                'Content-Type': 'application/json',
            }

            try:
                response = http_requests.post(url, json=payload, headers=headers)
                if response.status_code == 200:
                    data = response.json()
                    # Store pidx in session or order (if you have a field for it)
                    # Let's use the session or just pass it in return_url if needed, 
                    # but Khalti sends it back in GET params.
                    request.session['pending_order_pidx'] = data.get('pidx')
                    request.session['pending_order_number'] = order.order_number
                    return redirect(data.get('payment_url'))
                else:
                    messages.error(request, f"Khalti initiation failed: {response.text}")
                    return redirect('products:order_detail', order_number=order.order_number)
            except Exception as e:
                messages.error(request, f"Payment error: {str(e)}")
                return redirect('products:order_detail', order_number=order.order_number)
    
    context = {
        'cart_items': cart_items,
        'total': total,
    }
    return render(request, 'products/checkout.html', context)


@login_required
def order_detail(request, order_number):
    """View order details"""
    from .models import Order
    
    order = get_object_or_404(
        Order.objects.prefetch_related('items__product'),
        order_number=order_number,
        user=request.user
    )
    
    context = {
        'order': order,
    }
    return render(request, 'products/order_detail.html', context)


@login_required
def order_list(request):
    """View all user orders"""
    from .models import Order
    
    orders = Order.objects.filter(user=request.user).prefetch_related('items__product')
    
    context = {
        'orders': orders,
    }
    return render(request, 'products/order_list.html', context)


@login_required
def order_cancel(request, order_number):
    """Allow user to cancel a pending order"""
    order = get_object_or_404(Order, order_number=order_number, user=request.user)
    
    if order.status != 'pending':
        messages.error(request, f"Order #{order_number} cannot be cancelled as it's already {order.status}.")
        return redirect('products:order_detail', order_number=order_number)
    
    if request.method == 'POST':
        # Restore stock before cancelling
        for item in order.items.all():
            item.product.stock_quantity += item.quantity
            item.product.save()
            
        order.status = 'cancelled'
        order.save()
        messages.success(request, f"Order #{order_number} has been cancelled successfully.")
        return redirect('products:order_list')
        
    return render(request, 'products/order_confirm_cancel.html', {'order': order})


@login_required
def order_edit(request, order_number):
    """Allow user to edit shipping details of a pending order"""
    from .forms import OrderUpdateForm
    order = get_object_or_404(Order, order_number=order_number, user=request.user)
    
    if order.status != 'pending':
        messages.error(request, f"Order #{order_number} cannot be edited as it's already {order.status}.")
        return redirect('products:order_detail', order_number=order_number)
        
    if request.method == 'POST':
        form = OrderUpdateForm(request.POST, instance=order)
        if form.is_valid():
            form.save()
            messages.success(request, f"Order #{order_number} shipping details updated!")
            return redirect('products:order_detail', order_number=order_number)
    else:
        form = OrderUpdateForm(instance=order)
        
    return render(request, 'products/order_form_user.html', {
        'form': form, 
        'order': order,
        'action': 'Update Shipping Info'
    })


@login_required
def khalti_verify(request):
    """
    KPG-2 Callback: Handles return from Khalti and verifies status using lookup API.
    """
    import requests as http_requests
    from .models import Order, Payment
    from django.conf import settings
    
    pidx = request.GET.get('pidx')
    if not pidx:
        messages.error(request, "Invalid payment callback.")
        return redirect('products:cart')

    # Get order from session or verify against returned data
    order_number = request.session.get('pending_order_number')
    if not order_number:
        # Fallback/Security: Check if we can find an order with this pidx if stored
        messages.error(request, "Session expired or order not found.")
        return redirect('products:cart')

    order = get_object_or_404(Order, order_number=order_number, user=request.user)

    # Call lookup to verify
    url = f"{settings.KHALTI_BASE_URL}epayment/lookup/"
    payload = {"pidx": pidx}
    headers = {"Authorization": f"Key {settings.KHALTI_SECRET_KEY}"}

    try:
        response = http_requests.post(url, json=payload, headers=headers)
        if response.status_code == 200:
            data = response.json()
            if data.get('status') == 'Completed':
                order.payment_status = 'paid'
                order.payment_method = 'khalti'
                order.save()
                
                # Create payment record
                Payment.objects.create(
                    order=order,
                    user=request.user,
                    transaction_id=pidx,
                    payment_method='khalti',
                    amount=order.total_amount,
                    status='completed',
                    payment_response=data
                )
                
                # Send email confirmation
                from django.core.mail import send_mail
                
                subject = f"Payment Received: Order #{order.order_number}"
                
                context = {
                    'user_name': request.user.get_full_name() or request.user.username,
                    'order': order,
                    'site_url': request.build_absolute_uri('/')[:-1]
                }
                
                html_message = render_to_string('emails/order_confirmation.html', context)
                plain_message = strip_tags(html_message)
                
                try:
                    send_mail(
                        subject, 
                        plain_message, 
                        settings.DEFAULT_FROM_EMAIL, 
                        [request.user.email], 
                        html_message=html_message,
                        fail_silently=True
                    )
                except Exception as e:
                    print(f"DEBUG: Email confirmation error (Khalti): {e}")
                    pass
                
                # Cleanup session
                if 'pending_order_number' in request.session: del request.session['pending_order_number']
                if 'pending_order_pidx' in request.session: del request.session['pending_order_pidx']
                
                messages.success(request, f"Payment successful! Your order #{order.order_number} is processing.")
                return redirect('products:order_detail', order_number=order.order_number)
            else:
                messages.warning(request, f"Payment status: {data.get('status')}")
                return redirect('products:order_detail', order_number=order.order_number)
        else:
            messages.error(request, "Verification failed on Khalti's server.")
            return redirect('products:cart')
    except Exception as e:
        messages.error(request, f"Verification error: {str(e)}")
        return redirect('products:cart')


# Admin Order Management Views
@login_required
@admin_required
def admin_order_list(request):
    """Admin view to list all orders"""
    from django.db.models import Q
    from django.core.paginator import Paginator
    
    # Get filter parameters
    status_filter = request.GET.get('status', '')
    payment_status_filter = request.GET.get('payment_status', '')
    payment_method_filter = request.GET.get('payment_method', '')
    search_query = request.GET.get('search', '')
    
    orders = Order.objects.all().select_related('user').prefetch_related('items__product').order_by('-created_at')
    
    # Apply filters
    if status_filter:
        orders = orders.filter(status=status_filter)
    
    if payment_status_filter:
        orders = orders.filter(payment_status=payment_status_filter)
    
    if payment_method_filter:
        orders = orders.filter(payment_method=payment_method_filter)
    
    if search_query:
        orders = orders.filter(
            Q(order_number__icontains=search_query) |
            Q(user__username__icontains=search_query) |
            Q(user__email__icontains=search_query) |
            Q(user__first_name__icontains=search_query) |
            Q(user__last_name__icontains=search_query)
        )
    
    # Pagination - 10 orders per page
    paginator = Paginator(orders, 10)
    page_number = request.GET.get('page')
    page_obj = paginator.get_page(page_number)
    
    # Calculate statistics for all filtered orders (not just current page)
    total_orders = orders.count()
    pending_orders = orders.filter(status='pending').count()
    total_revenue = sum(order.total_amount for order in orders.filter(payment_status='paid'))
    pending_revenue = sum(order.total_amount for order in orders.filter(payment_method='cod', status__in=['pending', 'processing']))
    completed_orders = orders.filter(status='completed').count()
    
    context = {
        'page_obj': page_obj,
        'status_filter': status_filter,
        'payment_status_filter': payment_status_filter,
        'payment_method_filter': payment_method_filter,
        'search_query': search_query,
        'total_orders': total_orders,
        'pending_orders': pending_orders,
        'total_revenue': total_revenue,
        'pending_revenue': pending_revenue,
        'completed_orders': completed_orders,
    }
    return render(request, 'products/admin_order_list.html', context)


@login_required
@admin_required
def admin_order_detail(request, order_number):
    """Admin view to see order details"""
    order = get_object_or_404(Order, order_number=order_number)
    
    context = {
        'order': order,
    }
    return render(request, 'products/admin_order_detail.html', context)


@login_required
@admin_required
def update_order_status(request, order_number):
    """Update order status and payment status"""
    order = get_object_or_404(Order, order_number=order_number)
    
    if request.method == 'POST':
        new_status = request.POST.get('status')
        new_payment_status = request.POST.get('payment_status')
        status_changed = False
        
        if new_status in dict(Order.STATUS_CHOICES):
            order.status = new_status
            status_changed = True
        
        if new_payment_status in dict(Order.PAYMENT_STATUS_CHOICES):
            order.payment_status = new_payment_status
            status_changed = True
            
            # Sync corresponding Payment object
            from .models import Payment
            payment = Payment.objects.filter(order=order).first()
            if payment:
                if new_payment_status == 'paid' and payment.status != 'completed':
                    payment.status = 'completed'
                    payment.save()
                elif new_payment_status == 'pending' and payment.status != 'pending':
                    payment.status = 'pending'
                    payment.save()
                elif new_payment_status == 'failed' and payment.status != 'failed':
                    payment.status = 'failed'
                    payment.save()
            elif new_payment_status == 'paid':
                # If no payment object exists (e.g. COD) but admin marks as paid, create one
                from django.utils import timezone
                Payment.objects.create(
                    order=order,
                    user=order.user,
                    transaction_id=f"SHOP-ORDER-{order.id}-{timezone.now().strftime('%Y%m%d%H%M%S')}",
                    payment_method=order.payment_method or 'cod',
                    amount=order.total_amount,
                    status='completed',
                    payment_response={'source': 'Admin Status Update', 'recorded_by': request.user.username}
                )
        
        if status_changed:
            order.save()
            messages.success(request, f'Order #{order.order_number} updated successfully!')
        
        return redirect('products:admin_order_detail', order_number=order.order_number)
    
    return redirect('products:admin_order_list')


@login_required
@admin_required
def admin_order_refund(request, order_number):
    """Admin manually records a refund for an e-commerce order."""
    order = get_object_or_404(Order, order_number=order_number)
    
    if request.method == 'POST':
        try:
            refund_amount = decimal.Decimal(request.POST.get('refund_amount', '0'))
            refund_method = request.POST.get('refund_method', 'cash')
            notes = request.POST.get('notes', '')
            
            if refund_amount <= 0:
                messages.error(request, 'Refund amount must be greater than zero.')
                return redirect('products:admin_order_detail', order_number=order_number)
            
            if refund_amount > order.total_amount:
                messages.error(request, f'Refund amount cannot exceed order total (Rs. {order.total_amount}).')
                return redirect('products:admin_order_detail', order_number=order_number)
            
            # 1. Update order status and payment status
            order.status = 'refunded'
            order.payment_status = 'refunded'
            order.save()
            
            # 2. Record the refund as a Payment object with negative amount
            unique_id = f"REF-{uuid.uuid4().hex[:4]}-{timezone.now().strftime('%y%m%d')}"
            
            Payment.objects.create(
                order=order,
                user=order.user,
                transaction_id=f"PAY-{order.order_number}-{unique_id}",
                payment_method=refund_method,
                amount=-refund_amount,
                status='refunded',
                payment_response={
                    'source': 'Manual Order Refund',
                    'recorded_by': request.user.username,
                    'notes': notes,
                    'is_manual': True
                }
            )
            
            messages.success(request, f'Succesfully recorded refund of Rs. {refund_amount} for Order #{order.order_number}.')
            
        except (ValueError, decimal.InvalidOperation):
            messages.error(request, 'Invalid refund amount.')
            
        return redirect('products:admin_order_detail', order_number=order_number)

    return redirect('products:admin_order_detail', order_number=order_number)


# Admin Review Management Views
@login_required
@admin_required
def admin_review_list(request):
    """Admin view to list all product reviews"""
    from .models import Review
    from django.db.models import Q, Avg
    from django.core.paginator import Paginator
    
    # Get filter parameters
    rating_filter = request.GET.get('rating', '')
    product_filter = request.GET.get('product', '')
    search_query = request.GET.get('search', '')
    
    reviews = Review.objects.all().select_related('product', 'user').order_by('-created_at')
    
    # Apply filters
    if rating_filter:
        reviews = reviews.filter(rating=rating_filter)
    
    if product_filter:
        reviews = reviews.filter(product_id=product_filter)
    
    if search_query:
        reviews = reviews.filter(
            Q(user__username__icontains=search_query) |
            Q(user__email__icontains=search_query) |
            Q(product__name__icontains=search_query) |
            Q(comment__icontains=search_query)
        )
    
    # Pagination - 10 reviews per page
    paginator = Paginator(reviews, 10)
    page_number = request.GET.get('page')
    page_obj = paginator.get_page(page_number)
    
    # Get all products for filter dropdown
    products = Product.objects.filter(is_active=True).order_by('name')

    # Metrics
    total_reviews = Review.objects.count()
    approved_count = Review.objects.filter(is_approved=True).count()
    pending_count = Review.objects.filter(is_approved=False).count()
    avg_rating = Review.objects.aggregate(Avg('rating'))['rating__avg'] or 0
    five_star_count = Review.objects.filter(rating=5).count()
    
    context = {
        'page_obj': page_obj,
        'rating_filter': rating_filter,
        'product_filter': product_filter,
        'product_filter_int': int(product_filter) if product_filter else None,
        'search_query': search_query,
        'products': products,
        'total_reviews': total_reviews,
        'approved_count': approved_count,
        'pending_count': pending_count,
        'avg_rating': round(avg_rating, 1),
        'five_star_count': five_star_count,
    }
    return render(request, 'products/admin_review_list.html', context)


@login_required
@admin_required
def admin_approve_review(request, pk):
    """Toggle approval status of a product review"""
    from .models import Review
    review = get_object_or_404(Review, pk=pk)
    review.is_approved = not review.is_approved
    review.save()
    status = "approved" if review.is_approved else "unapproved"
    messages.success(request, f'Review has been {status}.')
    return redirect('products:admin_review_list')


@login_required
@admin_required
def admin_delete_review(request, pk):
    """Admin view to delete a review"""
    from .models import Review
    
    review = get_object_or_404(Review, pk=pk)
    
    if request.method == 'POST':
        product_name = review.product.name
        user_name = review.user.get_full_name() or review.user.username
        review.delete()
        messages.success(request, f'Review by {user_name} for {product_name} has been deleted.')
        return redirect('products:admin_review_list')
    
    return render(request, 'products/admin_review_confirm_delete.html', {'review': review})



# Payment Views
@login_required
def user_payment_list(request):
    """User view to see their payment history"""
    from .models import Payment
    from django.core.paginator import Paginator
    
    payments = Payment.objects.filter(user=request.user).select_related('order')
    
    # Pagination
    paginator = Paginator(payments, 10)
    page_number = request.GET.get('page')
    page_obj = paginator.get_page(page_number)
    
    context = {
        'page_obj': page_obj,
        'total_payments': payments.count(),
        'total_amount': sum(p.amount for p in payments if p.status == 'completed'),
    }
    return render(request, 'products/user_payment_list.html', context)


@login_required
@admin_required
def admin_payment_list(request):
    """Admin view to see all payments"""
    from .models import Payment
    from django.core.paginator import Paginator
    from django.db.models import Q
    
    # Get filter parameters
    status_filter = request.GET.get('status', '')
    payment_method_filter = request.GET.get('payment_method', '')
    search_query = request.GET.get('search', '')
    
    # Base queryset - include both order and appointment
    payments = Payment.objects.all().select_related('order', 'appointment', 'user').order_by('-created_at')
    
    # Apply filters
    if status_filter:
        payments = payments.filter(status=status_filter)
    
    if payment_method_filter:
        payments = payments.filter(payment_method=payment_method_filter)
    
    if search_query:
        payments = payments.filter(
            Q(transaction_id__icontains=search_query) |
            Q(user__username__icontains=search_query) |
            Q(user__email__icontains=search_query) |
            Q(order__order_number__icontains=search_query) |
            Q(appointment__id__icontains=search_query)
        )
    
    # Pagination
    paginator = Paginator(payments, 15)
    page_number = request.GET.get('page')
    page_obj = paginator.get_page(page_number)
    
    # Calculate statistics using DB aggregation for accuracy
    from django.db.models import Sum
    all_payments = Payment.objects.all()
    completed_payments_qs = all_payments.filter(status='completed')
    
    total_rev = completed_payments_qs.aggregate(total=Sum('amount'))['total'] or 0
    
    context = {
        'page_obj': page_obj,
        'status_filter': status_filter,
        'payment_method_filter': payment_method_filter,
        'search_query': search_query,
        'total_payments': all_payments.count(),
        'completed_payments': completed_payments_qs.count(),
        'total_revenue': total_rev,
        'status_choices': Payment.STATUS_CHOICES,
        'payment_method_choices': Payment.PAYMENT_METHOD_CHOICES,
    }
    return render(request, 'products/admin_payment_list.html', context)


@login_required
@admin_required
def admin_payment_detail(request, pk):
    """Admin view to see payment details"""
    from .models import Payment
    
    payment = get_object_or_404(Payment.objects.select_related('order', 'user'), pk=pk)
    
    context = {
        'payment': payment,
    }
    return render(request, 'products/admin_payment_detail.html', context)


@login_required
@admin_required
def export_orders_excel(request):
    """Export all or filtered orders to Excel"""
    import openpyxl
    from django.http import HttpResponse
    from django.utils import timezone
    from openpyxl.styles import Font, Alignment, PatternFill
    
    # Get filters from request
    status_filter = request.GET.get('status', '')
    payment_status_filter = request.GET.get('payment_status', '')
    search_query = request.GET.get('search', '')
    
    orders = Order.objects.all().select_related('user').order_by('-created_at')
    
    # Apply filters (same as admin_order_list)
    if status_filter:
        orders = orders.filter(status=status_filter)
    if payment_status_filter:
        orders = orders.filter(payment_status=payment_status_filter)
    if search_query:
        orders = orders.filter(
            Q(order_number__icontains=search_query) |
            Q(user__username__icontains=search_query) |
            Q(user__email__icontains=search_query)
        )
    
    # Create workbook
    wb = openpyxl.Workbook()
    ws = wb.active
    ws.title = "Orders Report"
    
    # Header cells
    headers = ['Order #', 'Customer', 'Email', 'Date', 'Amount (Rs.)', 'Method', 'Payment Status', 'Order Status']
    ws.append(headers)
    
    # Style headers
    header_font = Font(bold=True, color="FFFFFF")
    header_fill = PatternFill(start_color="4E73DF", end_color="4E73DF", fill_type="solid")
    for cell in ws[1]:
        cell.font = header_font
        cell.fill = header_fill
        cell.alignment = Alignment(horizontal='center')
    
    # Add data
    for order in orders:
        ws.append([
            order.order_number,
            order.user.get_full_name() or order.user.username,
            order.user.email,
            order.created_at.strftime('%Y-%m-%d %H:%M'),
            float(order.total_amount),
            order.payment_method.upper() if order.payment_method else 'COD',
            order.payment_status.upper(),
            order.status.upper()
        ])
    
    # Adjust column widths
    for column_cells in ws.columns:
        length = max(len(str(cell.value)) for cell in column_cells)
        ws.column_dimensions[column_cells[0].column_letter].width = length + 2

    # Prepare response
    timestamp = timezone.now().strftime('%Y%m%d_%H%M%S')
    response = HttpResponse(content_type='application/vnd.openxmlformats-officedocument.spreadsheetml.sheet')
    response['Content-Disposition'] = f'attachment; filename="Order_Report_{timestamp}.xlsx"'
    
    wb.save(response)
    return response


@login_required
@admin_required
def export_payments_excel(request):
    """Export all or filtered payments to Excel"""
    import openpyxl
    from django.http import HttpResponse
    from django.utils import timezone
    from openpyxl.styles import Font, Alignment, PatternFill
    
    # Get filters
    status_filter = request.GET.get('status', '')
    method_filter = request.GET.get('payment_method', '')
    
    payments = Payment.objects.all().select_related('user', 'order', 'appointment').order_by('-created_at')
    
    if status_filter:
        payments = payments.filter(status=status_filter)
    if method_filter:
        payments = payments.filter(payment_method=method_filter)
        
    wb = openpyxl.Workbook()
    ws = wb.active
    ws.title = "Payments Report"
    
    headers = ['Transaction ID', 'User', 'Email', 'Amount (Rs.)', 'Method', 'ID (Order/Appt)', 'Status', 'Date']
    ws.append(headers)
    
    header_font = Font(bold=True, color="FFFFFF")
    header_fill = PatternFill(start_color="1CC88A", end_color="1CC88A", fill_type="solid")
    for cell in ws[1]:
        cell.font = header_font
        cell.fill = header_fill
        cell.alignment = Alignment(horizontal='center')
        
    for p in payments:
        target_id = f"Order: {p.order.order_number}" if p.order else (f"Appt: {p.appointment.id}" if p.appointment else "N/A")
        ws.append([
            p.transaction_id,
            p.user.get_full_name() or p.user.username,
            p.user.email,
            float(p.amount),
            p.payment_method.upper() if p.payment_method else 'N/A',
            target_id,
            p.status.upper(),
            p.created_at.strftime('%Y-%m-%d %H:%M')
        ])
        
    for column_cells in ws.columns:
        length = max(len(str(cell.value)) for cell in column_cells if cell.value)
        ws.column_dimensions[column_cells[0].column_letter].width = length + 2

    timestamp = timezone.now().strftime('%Y%m%d_%H%M%S')
    response = HttpResponse(content_type='application/vnd.openxmlformats-officedocument.spreadsheetml.sheet')
    response['Content-Disposition'] = f'attachment; filename="Payment_Report_{timestamp}.xlsx"'
    wb.save(response)
    return response


@login_required
@admin_required
def export_orders_csv(request):
    """Export all or filtered orders to CSV"""
    from django.http import HttpResponse
    from django.utils import timezone
    import csv

    # Get filters from request
    status_filter = request.GET.get('status', '')
    payment_status_filter = request.GET.get('payment_status', '')
    search_query = request.GET.get('search', '')

    orders = Order.objects.all().select_related('user').order_by('-created_at')

    # Apply filters (same as admin_order_list)
    if status_filter:
        orders = orders.filter(status=status_filter)
    if payment_status_filter:
        orders = orders.filter(payment_status=payment_status_filter)
    if search_query:
        orders = orders.filter(
            Q(order_number__icontains=search_query) |
            Q(user__username__icontains=search_query) |
            Q(user__email__icontains=search_query)
        )

    # Prepare response
    timestamp = timezone.now().strftime('%Y%m%d_%H%M%S')
    response = HttpResponse(content_type='text/csv')
    response['Content-Disposition'] = f'attachment; filename="Order_Report_{timestamp}.csv"'

    writer = csv.writer(response)
    writer.writerow(['Order #', 'Customer', 'Email', 'Date', 'Amount (Rs.)', 'Method', 'Payment Status', 'Order Status'])

    for order in orders:
        writer.writerow([
            order.order_number,
            order.user.get_full_name() or order.user.username,
            order.user.email,
            order.created_at.strftime('%Y-%m-%d %H:%M'),
            float(order.total_amount),
            order.payment_method.upper() if order.payment_method else 'COD',
            order.payment_status.upper(),
            order.status.upper()
        ])

    return response


@login_required
@admin_required
def export_payments_csv(request):
    """Export all or filtered payments to CSV"""
    from django.http import HttpResponse
    from django.utils import timezone
    import csv

    # Get filters
    status_filter = request.GET.get('status', '')
    method_filter = request.GET.get('payment_method', '')

    payments = Payment.objects.all().select_related('user', 'order', 'appointment').order_by('-created_at')

    if status_filter:
        payments = payments.filter(status=status_filter)
    if method_filter:
        payments = payments.filter(payment_method=method_filter)

    # Prepare response
    timestamp = timezone.now().strftime('%Y%m%d_%H%M%S')
    response = HttpResponse(content_type='text/csv')
    response['Content-Disposition'] = f'attachment; filename="Payment_Report_{timestamp}.csv"'

    writer = csv.writer(response)
    writer.writerow(['Transaction ID', 'User', 'Email', 'Amount (Rs.)', 'Method', 'ID (Order/Appt)', 'Status', 'Date'])

    for p in payments:
        target_id = f"Order: {p.order.order_number}" if p.order else (f"Appt: {p.appointment.id}" if p.appointment else "N/A")
        writer.writerow([
            p.transaction_id,
            p.user.get_full_name() or p.user.username,
            p.user.email,
            float(p.amount),
            p.payment_method.upper() if p.payment_method else 'N/A',
            target_id,
            p.status.upper(),
            p.created_at.strftime('%Y-%m-%d %H:%M')
        ])

    return response


@login_required
@admin_required
def export_sales_csv(request):
    """Redirect to accounts version of sales report"""
    from django.shortcuts import redirect
    from django.urls import reverse
    url = reverse('export_sales_csv') + f"?month={request.GET.get('month', '')}&year={request.GET.get('year', '')}"
    return redirect(url)


